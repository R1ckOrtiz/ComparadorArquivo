from __future__ import annotations

import csv
import difflib
import hashlib
import io
import json
import mimetypes
import os
import re
import tempfile
import zipfile
from dataclasses import dataclass, field
from itertools import zip_longest
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


DOCX_NAMESPACES = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
TEXT_EXTENSIONS = {
    ".txt",
    ".md",
    ".csv",
    ".tsv",
    ".json",
    ".xml",
    ".html",
    ".htm",
    ".svg",
    ".yaml",
    ".yml",
    ".ini",
    ".cfg",
    ".log",
    ".sql",
    ".py",
    ".js",
    ".ts",
    ".tsx",
    ".jsx",
    ".css",
    ".scss",
    ".less",
    ".java",
    ".cs",
    ".go",
    ".rs",
    ".rb",
    ".php",
    ".sh",
    ".bat",
    ".ps1",
}
JSON_EXTENSIONS = {".json"}
CSV_EXTENSIONS = {".csv", ".tsv"}
PDF_EXTENSIONS = {".pdf"}
DOCX_EXTENSIONS = {".docx"}
XLSX_EXTENSIONS = {".xlsx", ".xlsm"}
XML_EXTENSIONS = {".xml", ".svg"}
MAX_REPORTED_BINARY_RANGES = 5
MAX_DISPLAY_CHANGES = 40
MAX_DIFF_LINES = 250
MAX_ROW_CONTEXT_CELLS = 8


@dataclass
class DocumentProfile:
    filename: str
    source_type: str
    format_label: str
    mode: str
    metadata: dict[str, Any]
    warnings: list[str] = field(default_factory=list)
    lines: list[str] = field(default_factory=list)
    records: dict[str, str] = field(default_factory=dict)
    entries: dict[str, dict[str, Any]] = field(default_factory=dict)


def compare_documents(
    filename_a: str,
    content_a: bytes,
    filename_b: str,
    content_b: bytes,
) -> dict[str, Any]:
    profile_a = inspect_document(filename_a, content_a)
    profile_b = inspect_document(filename_b, content_b)

    if profile_a.mode == "structured" and profile_b.mode == "structured":
        result = compare_structured_profiles(profile_a, profile_b)
    elif profile_a.mode == "excel" and profile_b.mode == "excel":
        result = compare_excel_profiles(profile_a, content_a, profile_b, content_b)
    elif is_textual_profile(profile_a) and is_textual_profile(profile_b):
        result = compare_text_profiles(profile_a, profile_b)
    elif profile_a.mode == "archive" and profile_b.mode == "archive":
        result = compare_archive_profiles(profile_a, profile_b)
    else:
        result = compare_binary_profiles(profile_a, content_a, profile_b, content_b)

    result["left"] = build_profile_summary(profile_a)
    result["right"] = build_profile_summary(profile_b)
    result["summary_items"] = build_summary_items(result["mode"], result["summary"])
    result["highlight_items"] = build_highlight_items(result["mode"], result["summary"])
    result["overview"] = build_comparison_overview(
        result["mode"],
        result["summary"],
        result["identical"],
        result["changes"],
    )
    result["focus_points"] = build_focus_points(result["changes"])
    result["warnings"] = profile_a.warnings + profile_b.warnings + result.pop("notes", [])
    return result


def compare_uploaded_files(file_a: Any, file_b: Any) -> dict[str, Any]:
    filename_a = file_a.filename
    filename_b = file_b.filename
    extension_a = Path(filename_a).suffix.lower()
    extension_b = Path(filename_b).suffix.lower()

    if extension_a in XLSX_EXTENSIONS and extension_b in XLSX_EXTENSIONS:
        return compare_excel_uploads(file_a, file_b)

    return compare_documents(filename_a, file_a.read(), filename_b, file_b.read())


def compare_excel_uploads(file_a: Any, file_b: Any) -> dict[str, Any]:
    path_a = materialize_upload_to_tempfile(file_a)
    path_b = materialize_upload_to_tempfile(file_b)

    try:
        profile_a = build_excel_profile_from_path(file_a.filename, path_a)
        profile_b = build_excel_profile_from_path(file_b.filename, path_b)
        result = compare_excel_profiles(profile_a, path_a, profile_b, path_b)
        result["left"] = build_profile_summary(profile_a)
        result["right"] = build_profile_summary(profile_b)
        result["summary_items"] = build_summary_items(result["mode"], result["summary"])
        result["highlight_items"] = build_highlight_items(result["mode"], result["summary"])
        result["overview"] = build_comparison_overview(
            result["mode"],
            result["summary"],
            result["identical"],
            result["changes"],
        )
        result["focus_points"] = build_focus_points(result["changes"])
        result["warnings"] = profile_a.warnings + profile_b.warnings + result.pop("notes", [])
        return result
    finally:
        cleanup_temp_path(path_a)
        cleanup_temp_path(path_b)


def inspect_document(filename: str, content: bytes) -> DocumentProfile:
    extension = Path(filename).suffix.lower()
    mime_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    metadata = build_metadata(content, mime_type, extension)
    warnings: list[str] = []

    if extension in PDF_EXTENSIONS:
        try:
            return DocumentProfile(
                filename=filename,
                source_type="pdf",
                format_label="PDF",
                mode="text",
                metadata=metadata,
                lines=extract_pdf_lines(content),
            )
        except Exception as exc:  # pragma: no cover
            warnings.append(f"Falha ao extrair texto do PDF: {exc}")

    if extension in DOCX_EXTENSIONS:
        try:
            return DocumentProfile(
                filename=filename,
                source_type="docx",
                format_label="DOCX",
                mode="text",
                metadata=metadata,
                lines=extract_docx_lines(content),
            )
        except Exception as exc:  # pragma: no cover
            warnings.append(f"Falha ao extrair texto do DOCX: {exc}")

    if extension in XLSX_EXTENSIONS:
        try:
            return DocumentProfile(
                filename=filename,
                source_type="xlsx",
                format_label="XLSX",
                mode="excel",
                metadata=metadata,
                warnings=warnings,
            )
        except Exception as exc:  # pragma: no cover
            warnings.append(f"Falha ao extrair conteúdo do XLSX: {exc}")

    text_hint = extension in TEXT_EXTENSIONS or mime_type.startswith("text/")
    decoded_text = decode_text_content(content, allow_latin1=text_hint)

    if extension in JSON_EXTENSIONS and decoded_text is not None:
        try:
            return DocumentProfile(
                filename=filename,
                source_type="json",
                format_label="JSON",
                mode="structured",
                metadata=metadata,
                warnings=warnings,
                records=normalize_json_records(decoded_text),
            )
        except json.JSONDecodeError:
            warnings.append("JSON inválido; comparação feita como texto simples.")

    if extension in CSV_EXTENSIONS and decoded_text is not None:
        return DocumentProfile(
            filename=filename,
            source_type="csv",
            format_label="CSV/TSV",
            mode="structured",
            metadata=metadata,
            warnings=warnings,
            records=normalize_csv_records(decoded_text),
        )

    if extension in XML_EXTENSIONS and decoded_text is not None:
        try:
            return DocumentProfile(
                filename=filename,
                source_type="xml",
                format_label="XML/SVG",
                mode="structured",
                metadata=metadata,
                warnings=warnings,
                records=normalize_xml_records(decoded_text),
            )
        except ET.ParseError:
            warnings.append("XML/SVG inválido; comparação feita como texto simples.")

    if decoded_text is not None and looks_like_json(decoded_text):
        try:
            return DocumentProfile(
                filename=filename,
                source_type="json",
                format_label="JSON",
                mode="structured",
                metadata=metadata,
                warnings=warnings,
                records=normalize_json_records(decoded_text),
            )
        except json.JSONDecodeError:
            pass

    if decoded_text is not None:
        return DocumentProfile(
            filename=filename,
            source_type="text",
            format_label="Texto legível",
            mode="text",
            metadata=metadata,
            warnings=warnings,
            lines=normalize_plain_text_lines(decoded_text),
        )

    if is_zip_archive(content):
        try:
            return DocumentProfile(
                filename=filename,
                source_type="archive",
                format_label="Estrutura ZIP",
                mode="archive",
                metadata=metadata,
                warnings=warnings,
                entries=inspect_zip_entries(content),
            )
        except Exception as exc:  # pragma: no cover
            warnings.append(f"Falha ao inspecionar a estrutura ZIP: {exc}")

    return DocumentProfile(
        filename=filename,
        source_type="binary",
        format_label="Binário genérico",
        mode="binary",
        metadata=metadata,
        warnings=warnings,
    )


def compare_text_profiles(profile_a: DocumentProfile, profile_b: DocumentProfile) -> dict[str, Any]:
    lines_a = profile_to_lines(profile_a)
    lines_b = profile_to_lines(profile_b)
    matcher = build_matcher(lines_a, lines_b)
    changes: list[dict[str, Any]] = []
    summary = {"added": 0, "removed": 0, "changed": 0, "unchanged": 0}

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            summary["unchanged"] += i2 - i1
            continue

        if tag == "replace":
            append_refined_text_changes(
                changes,
                summary,
                lines_a[i1:i2],
                lines_b[j1:j2],
                i1,
                j1,
            )
            continue

        changes.append(build_text_change(tag, lines_a[i1:i2], lines_b[j1:j2], i1, i2, j1, j2))
        update_text_summary(summary, tag, i1, i2, j1, j2)

    diff_lines = list(
        difflib.unified_diff(
            lines_a,
            lines_b,
            fromfile=profile_a.filename,
            tofile=profile_b.filename,
            lineterm="",
        )
    )

    limited_changes, change_notes = limit_changes(changes)
    limited_diff, diff_notes = limit_diff_text(diff_lines)
    text_similarity = format_percent(compute_similarity("\n".join(lines_a), "\n".join(lines_b)))

    return {
        "mode": "text",
        "mode_label": "Comparação detalhada",
        "strategy_label": build_text_strategy_label(profile_a, profile_b),
        "identical": not changes,
        "summary": {
            **summary,
            "differences": len(changes),
            "lines_a": len(lines_a),
            "lines_b": len(lines_b),
            "similarity": text_similarity,
        },
        "changes": limited_changes,
        "unified_diff": limited_diff,
        "notes": change_notes + diff_notes,
    }


def compare_structured_profiles(profile_a: DocumentProfile, profile_b: DocumentProfile) -> dict[str, Any]:
    changes: list[dict[str, Any]] = []
    summary = {"added": 0, "removed": 0, "changed": 0, "unchanged": 0}
    paths = sorted(set(profile_a.records) | set(profile_b.records))

    for path in paths:
        location_label = format_structure_location(path)

        if path not in profile_a.records:
            summary["added"] += 1
            changes.append(
                {
                    "type": "Adicionado",
                    "path": path,
                    "focus_label": location_label,
                    "location_label": location_label,
                    "before_value": None,
                    "after_value": profile_b.records[path],
                    "before_profile": None,
                    "after_profile": describe_value_profile(profile_b.records[path]),
                    "details": f"{location_label} foi adicionado no segundo arquivo.",
                }
            )
            continue

        if path not in profile_b.records:
            summary["removed"] += 1
            changes.append(
                {
                    "type": "Removido",
                    "path": path,
                    "focus_label": location_label,
                    "location_label": location_label,
                    "before_value": profile_a.records[path],
                    "after_value": None,
                    "before_profile": describe_value_profile(profile_a.records[path]),
                    "after_profile": None,
                    "details": f"{location_label} foi removido no segundo arquivo.",
                }
            )
            continue

        if profile_a.records[path] != profile_b.records[path]:
            summary["changed"] += 1
            changes.append(
                {
                    "type": "Alterado",
                    "path": path,
                    "focus_label": location_label,
                    "location_label": location_label,
                    "before_value": profile_a.records[path],
                    "after_value": profile_b.records[path],
                    "before_profile": describe_value_profile(profile_a.records[path]),
                    "after_profile": describe_value_profile(profile_b.records[path]),
                    "inline_diff": build_inline_diff_text(profile_a.records[path], profile_b.records[path]),
                    "similarity": format_percent(
                        compute_similarity(profile_a.records[path], profile_b.records[path])
                    ),
                    "details": f"{location_label} teve o valor alterado entre os arquivos.",
                }
            )
            continue

        summary["unchanged"] += 1

    diff_lines = list(
        difflib.unified_diff(
            render_record_lines(profile_a.records),
            render_record_lines(profile_b.records),
            fromfile=profile_a.filename,
            tofile=profile_b.filename,
            lineterm="",
        )
    )

    limited_changes, change_notes = limit_changes(changes)
    limited_diff, diff_notes = limit_diff_text(diff_lines)

    return {
        "mode": "structured",
        "mode_label": "Comparação semântica",
        "strategy_label": build_structured_strategy_label(profile_a, profile_b),
        "identical": not changes,
        "summary": {
            **summary,
            "differences": len(changes),
            "fields_a": len(profile_a.records),
            "fields_b": len(profile_b.records),
            "similarity": format_percent(
                compute_similarity(
                    "\n".join(render_record_lines(profile_a.records)),
                    "\n".join(render_record_lines(profile_b.records)),
                )
            ),
        },
        "changes": limited_changes,
        "unified_diff": limited_diff,
        "notes": change_notes + diff_notes,
    }


def compare_excel_profiles(
    profile_a: DocumentProfile,
    content_a: bytes | str,
    profile_b: DocumentProfile,
    content_b: bytes | str,
) -> dict[str, Any]:
    workbook_a = load_excel_workbook(content_a)
    workbook_b = load_excel_workbook(content_b)
    changes: list[dict[str, Any]] = []
    hidden_changes = 0
    affected_sheets: set[str] = set()
    summary = {
        "added": 0,
        "removed": 0,
        "changed": 0,
        "unchanged": 0,
        "rows_compared": 0,
        "rows_changed": 0,
        "sheets_added": 0,
        "sheets_removed": 0,
    }

    try:
        sheets_a = workbook_a.sheetnames
        sheets_b = workbook_b.sheetnames
        profile_a.metadata["sheet_count"] = len(sheets_a)
        profile_b.metadata["sheet_count"] = len(sheets_b)
        profile_a.metadata["sheet_names"] = list(sheets_a)
        profile_b.metadata["sheet_names"] = list(sheets_b)

        set_a = set(sheets_a)
        set_b = set(sheets_b)

        for sheet_name in sorted(set_a - set_b):
            summary["removed"] += 1
            summary["sheets_removed"] += 1
            hidden_changes += append_capped_change(
                changes,
                {
                    "type": "Removido",
                    "path": f"Planilha {sheet_name}",
                    "location_label": f"Planilha \"{sheet_name}\"",
                    "details": f"A planilha \"{sheet_name}\" existe apenas no primeiro arquivo.",
                },
            )

        for sheet_name in sorted(set_b - set_a):
            summary["added"] += 1
            summary["sheets_added"] += 1
            hidden_changes += append_capped_change(
                changes,
                {
                    "type": "Adicionado",
                    "path": f"Planilha {sheet_name}",
                    "location_label": f"Planilha \"{sheet_name}\"",
                    "details": f"A planilha \"{sheet_name}\" existe apenas no segundo arquivo.",
                },
            )

        common_sheets = [sheet_name for sheet_name in sheets_a if sheet_name in set_b]

        for sheet_name in common_sheets:
            worksheet_a = workbook_a[sheet_name]
            worksheet_b = workbook_b[sheet_name]
            row_iter_a = worksheet_a.iter_rows(values_only=True)
            row_iter_b = worksheet_b.iter_rows(values_only=True)
            header_row_a: list[Any] = []
            header_row_b: list[Any] = []

            for row_index, (row_a, row_b) in enumerate(zip_longest(row_iter_a, row_iter_b, fillvalue=()), start=1):
                normalized_row_a = trim_excel_row(row_a)
                normalized_row_b = trim_excel_row(row_b)

                if not normalized_row_a and not normalized_row_b:
                    continue

                if row_index == 1:
                    header_row_a = normalized_row_a
                    header_row_b = normalized_row_b

                summary["rows_compared"] += 1
                row_changed = False

                for col_index in range(max(len(normalized_row_a), len(normalized_row_b))):
                    before_value = normalize_excel_cell(
                        normalized_row_a[col_index] if col_index < len(normalized_row_a) else None
                    )
                    after_value = normalize_excel_cell(
                        normalized_row_b[col_index] if col_index < len(normalized_row_b) else None
                    )

                    if before_value == after_value:
                        if before_value is not None:
                            summary["unchanged"] += 1
                        continue

                    row_changed = True
                    affected_sheets.add(sheet_name)
                    column_letter = get_column_letter(col_index + 1)
                    cell_path = f"{sheet_name}!{column_letter}{row_index}"
                    location_label = (
                        f"Planilha \"{sheet_name}\", linha {row_index}, coluna {column_letter}"
                    )
                    should_capture_detail = len(changes) < MAX_DISPLAY_CHANGES
                    column_name = derive_excel_column_name(col_index, header_row_a, header_row_b)
                    row_reference = build_excel_row_reference(
                        normalized_row_a,
                        normalized_row_b,
                        header_row_a,
                        header_row_b,
                        row_index,
                    )
                    focus_label = build_excel_focus_label(sheet_name, row_index, column_letter, column_name, row_reference)
                    row_context_before = render_excel_row_context(normalized_row_a, header_row_a) if should_capture_detail else None
                    row_context_after = render_excel_row_context(normalized_row_b, header_row_b) if should_capture_detail else None
                    before_profile = describe_value_profile(before_value)
                    after_profile = describe_value_profile(after_value)

                    if before_value is None:
                        summary["added"] += 1
                        change = {
                            "type": "Adicionado",
                            "path": cell_path,
                            "location_label": location_label,
                            "focus_label": focus_label,
                            "sheet_name": sheet_name,
                            "row_number": row_index,
                            "column_letter": column_letter,
                            "column_name": column_name,
                            "row_reference": row_reference,
                            "before_value": None,
                            "after_value": after_value,
                            "before_profile": None,
                            "after_profile": after_profile,
                            "row_context_before": row_context_before,
                            "row_context_after": row_context_after,
                            "details": (
                                f"{focus_label} foi preenchido apenas no segundo arquivo."
                            ),
                        }
                    elif after_value is None:
                        summary["removed"] += 1
                        change = {
                            "type": "Removido",
                            "path": cell_path,
                            "location_label": location_label,
                            "focus_label": focus_label,
                            "sheet_name": sheet_name,
                            "row_number": row_index,
                            "column_letter": column_letter,
                            "column_name": column_name,
                            "row_reference": row_reference,
                            "before_value": before_value,
                            "after_value": None,
                            "before_profile": before_profile,
                            "after_profile": None,
                            "row_context_before": row_context_before,
                            "row_context_after": row_context_after,
                            "details": (
                                f"{focus_label} foi removido no segundo arquivo."
                            ),
                        }
                    else:
                        summary["changed"] += 1
                        change = {
                            "type": "Alterado",
                            "path": cell_path,
                            "location_label": location_label,
                            "focus_label": focus_label,
                            "sheet_name": sheet_name,
                            "row_number": row_index,
                            "column_letter": column_letter,
                            "column_name": column_name,
                            "row_reference": row_reference,
                            "before_value": before_value,
                            "after_value": after_value,
                            "before_profile": before_profile,
                            "after_profile": after_profile,
                            "row_context_before": row_context_before,
                            "row_context_after": row_context_after,
                            "inline_diff": build_inline_diff_text(before_value, after_value) if should_capture_detail else None,
                            "similarity": format_percent(compute_similarity(before_value, after_value)),
                            "details": (
                                f"{focus_label} mudou de \"{truncate_text(before_value, 72)}\" para "
                                f"\"{truncate_text(after_value, 72)}\"."
                            ),
                        }

                    hidden_changes += append_capped_change(changes, change)

                if row_changed:
                    summary["rows_changed"] += 1

        notes = []
        if hidden_changes:
            notes.append(
                f"Exibindo apenas as primeiras {MAX_DISPLAY_CHANGES} alterações de {len(changes) + hidden_changes}."
            )

        total_items = summary["added"] + summary["removed"] + summary["changed"] + summary["unchanged"]
        similarity = 1.0 if total_items == 0 else summary["unchanged"] / total_items

        return {
            "mode": "excel",
            "mode_label": "Comparação de planilha",
            "strategy_label": "Comparação otimizada por planilha, linha e célula",
            "identical": summary["added"] == 0 and summary["removed"] == 0 and summary["changed"] == 0,
            "summary": {
                **summary,
                "differences": summary["added"] + summary["removed"] + summary["changed"],
                "sheets_a": len(sheets_a),
                "sheets_b": len(sheets_b),
                "sheets_changed": len(affected_sheets),
                "similarity": format_percent(similarity),
            },
            "changes": changes,
            "unified_diff": None,
            "notes": notes,
        }
    finally:
        workbook_a.close()
        workbook_b.close()


def compare_archive_profiles(profile_a: DocumentProfile, profile_b: DocumentProfile) -> dict[str, Any]:
    changes: list[dict[str, Any]] = []
    summary = {"added": 0, "removed": 0, "changed": 0}
    all_entries = sorted(set(profile_a.entries) | set(profile_b.entries))

    for entry_name in all_entries:
        if entry_name not in profile_a.entries:
            summary["added"] += 1
            changes.append(
                {
                    "type": "Adicionado",
                    "path": entry_name,
                    "location_label": f"Entrada interna {entry_name}",
                    "details": "Essa entrada existe apenas no segundo arquivo.",
                }
            )
            continue

        if entry_name not in profile_b.entries:
            summary["removed"] += 1
            changes.append(
                {
                    "type": "Removido",
                    "path": entry_name,
                    "location_label": f"Entrada interna {entry_name}",
                    "details": "Essa entrada existe apenas no primeiro arquivo.",
                }
            )
            continue

        if profile_a.entries[entry_name] != profile_b.entries[entry_name]:
            summary["changed"] += 1
            changes.append(
                {
                    "type": "Alterado",
                    "path": entry_name,
                    "location_label": f"Entrada interna {entry_name}",
                    "details": (
                        f"Tamanho {profile_a.entries[entry_name]['file_size']} -> "
                        f"{profile_b.entries[entry_name]['file_size']}, comprimido "
                        f"{profile_a.entries[entry_name]['compressed_size']} -> "
                        f"{profile_b.entries[entry_name]['compressed_size']}, CRC "
                        f"{profile_a.entries[entry_name]['crc']} -> "
                        f"{profile_b.entries[entry_name]['crc']}"
                    ),
                }
            )

    limited_changes, change_notes = limit_changes(changes)

    return {
        "mode": "archive",
        "mode_label": "Comparação estrutural",
        "strategy_label": "Comparação por entradas internas do arquivo ZIP",
        "identical": not changes,
        "summary": {
            **summary,
            "differences": len(changes),
            "entries_a": len(profile_a.entries),
            "entries_b": len(profile_b.entries),
        },
        "changes": limited_changes,
        "unified_diff": None,
        "notes": change_notes,
    }


def compare_binary_profiles(
    profile_a: DocumentProfile,
    content_a: bytes,
    profile_b: DocumentProfile,
    content_b: bytes,
) -> dict[str, Any]:
    ranges = find_binary_difference_ranges(content_a, content_b)
    first_difference = ranges[0][0] if ranges else None
    different_bytes = sum(end - start for start, end in ranges)

    return {
        "mode": "binary",
        "mode_label": "Comparação binária",
        "strategy_label": "Análise por hash, tamanho e faixas divergentes",
        "identical": not ranges,
        "summary": {
            "size_a": len(content_a),
            "size_b": len(content_b),
            "size_delta": len(content_b) - len(content_a),
            "different_bytes": different_bytes,
            "difference_ranges": len(ranges),
            "first_difference_offset": first_difference,
        },
        "changes": build_binary_changes(ranges, content_a, content_b),
        "unified_diff": None,
    }


def append_refined_text_changes(
    changes: list[dict[str, Any]],
    summary: dict[str, int],
    before_lines: list[str],
    after_lines: list[str],
    before_offset: int,
    after_offset: int,
) -> None:
    submatcher = build_matcher(before_lines, after_lines)

    for tag, i1, i2, j1, j2 in submatcher.get_opcodes():
        if tag == "equal":
            summary["unchanged"] += i2 - i1
            continue

        abs_i1 = before_offset + i1
        abs_i2 = before_offset + i2
        abs_j1 = after_offset + j1
        abs_j2 = after_offset + j2

        if tag == "replace" and (i2 - i1) == 1 and (j2 - j1) == 1:
            before_line = before_lines[i1]
            after_line = after_lines[j1]
            summary["changed"] += 1
            changes.append(
                {
                    "type": "Alterado",
                    "focus_label": describe_text_location(abs_i1, abs_i2, abs_j1, abs_j2),
                    "before_range": human_range(abs_i1, abs_i2),
                    "after_range": human_range(abs_j1, abs_j2),
                    "location_label": describe_text_location(abs_i1, abs_i2, abs_j1, abs_j2),
                    "before_excerpt": [before_line],
                    "after_excerpt": [after_line],
                    "inline_diff": build_inline_diff_text(before_line, after_line),
                    "similarity": format_percent(compute_similarity(before_line, after_line)),
                    "details": describe_text_change("replace", abs_i1, abs_i2, abs_j1, abs_j2),
                }
            )
            continue

        changes.append(build_text_change(tag, before_lines[i1:i2], after_lines[j1:j2], abs_i1, abs_i2, abs_j1, abs_j2))
        update_text_summary(summary, tag, abs_i1, abs_i2, abs_j1, abs_j2)


def build_text_change(
    tag: str,
    before_slice: list[str],
    after_slice: list[str],
    i1: int,
    i2: int,
    j1: int,
    j2: int,
) -> dict[str, Any]:
    if tag == "insert":
        change_type = "Adicionado"
    elif tag == "delete":
        change_type = "Removido"
    else:
        change_type = "Alterado"

    change = {
        "type": change_type,
        "focus_label": describe_text_location(i1, i2, j1, j2),
        "before_range": human_range(i1, i2),
        "after_range": human_range(j1, j2),
        "location_label": describe_text_location(i1, i2, j1, j2),
        "before_excerpt": excerpt_lines(before_slice),
        "after_excerpt": excerpt_lines(after_slice),
        "details": describe_text_change(tag, i1, i2, j1, j2),
    }

    return change


def update_text_summary(
    summary: dict[str, int],
    tag: str,
    i1: int,
    i2: int,
    j1: int,
    j2: int,
) -> None:
    if tag == "insert":
        summary["added"] += j2 - j1
    elif tag == "delete":
        summary["removed"] += i2 - i1
    else:
        summary["changed"] += max(i2 - i1, j2 - j1)


def build_binary_changes(
    ranges: list[tuple[int, int]],
    content_a: bytes,
    content_b: bytes,
) -> list[dict[str, Any]]:
    if not ranges:
        return []

    changes: list[dict[str, Any]] = []

    for start, end in ranges[:MAX_REPORTED_BINARY_RANGES]:
        changes.append(
            {
                "type": "Alterado",
                "path": format_binary_range(start, end),
                "location_label": f"Faixa binária {format_binary_range(start, end)}",
                "before_value": render_binary_window(content_a, start, end),
                "after_value": render_binary_window(content_b, start, end),
                "details": f"Foram encontrados {end - start} byte(s) divergentes nesta faixa.",
            }
        )

    hidden_ranges = len(ranges) - MAX_REPORTED_BINARY_RANGES
    if hidden_ranges > 0:
        changes.append(
            {
                "type": "Informativo",
                "path": "Faixas adicionais",
                "location_label": "Faixas binárias adicionais",
                "details": f"Existem mais {hidden_ranges} faixa(s) divergentes não exibidas.",
            }
        )

    return changes


def find_binary_difference_ranges(content_a: bytes, content_b: bytes) -> list[tuple[int, int]]:
    ranges: list[tuple[int, int]] = []
    current_start: int | None = None
    max_length = max(len(content_a), len(content_b))

    for offset in range(max_length):
        byte_a = content_a[offset] if offset < len(content_a) else None
        byte_b = content_b[offset] if offset < len(content_b) else None

        if byte_a != byte_b:
            if current_start is None:
                current_start = offset
            continue

        if current_start is not None:
            ranges.append((current_start, offset))
            current_start = None

    if current_start is not None:
        ranges.append((current_start, max_length))

    return ranges


def render_binary_window(content: bytes, start: int, end: int) -> str:
    if start >= len(content):
        return "[sem bytes nesta faixa]"

    window_start = max(0, start - 8)
    window_end = min(len(content), end + 8)
    payload = content[window_start:window_end]
    return f"offsets {window_start}-{window_end - 1}: {payload.hex(' ')}"


def format_binary_range(start: int, end: int) -> str:
    last = end - 1
    if start == last:
        return f"Offset {start}"
    return f"Offsets {start}-{last}"


def append_capped_change(changes: list[dict[str, Any]], change: dict[str, Any]) -> int:
    if len(changes) < MAX_DISPLAY_CHANGES:
        changes.append(change)
        return 0

    return 1


def trim_excel_row(row: tuple[Any, ...] | list[Any]) -> list[Any]:
    values = list(row)

    while values and normalize_excel_cell(values[-1]) is None:
        values.pop()

    return values


def normalize_excel_cell(value: Any) -> str | None:
    if value is None:
        return None

    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else ""

    if hasattr(value, "isoformat"):
        return value.isoformat()

    return str(value)


def render_excel_row_context(row: list[Any], header_row: list[Any] | None = None) -> str:
    populated_cells: list[str] = []
    extra_cells = 0

    for col_index, raw_value in enumerate(row, start=1):
        cell_value = normalize_excel_cell(raw_value)
        if cell_value in (None, ""):
            continue

        header_name = derive_excel_column_name(col_index - 1, header_row or [], header_row or [])
        label = header_name or get_column_letter(col_index)
        if len(populated_cells) < MAX_ROW_CONTEXT_CELLS:
            populated_cells.append(
                f"{label}={truncate_text(cell_value, 48)}"
            )
        else:
            extra_cells += 1

    if not populated_cells:
        return "[linha sem conteúdo relevante]"

    if extra_cells:
        populated_cells.append(f"... (+{extra_cells} célula(s))")

    return " | ".join(populated_cells)


def truncate_text(value: str, limit: int) -> str:
    if len(value) <= limit:
        return value
    return f"{value[:limit]}..."


def derive_excel_column_name(col_index: int, header_row_a: list[Any], header_row_b: list[Any]) -> str | None:
    for header_row in (header_row_a, header_row_b):
        if col_index < len(header_row):
            header_value = normalize_excel_cell(header_row[col_index])
            if header_value not in (None, ""):
                return header_value
    return None


def build_excel_row_reference(
    row_a: list[Any],
    row_b: list[Any],
    header_row_a: list[Any],
    header_row_b: list[Any],
    row_index: int,
) -> str | None:
    if row_index == 1:
        return "linha de cabeçalho"

    source_row = row_b if any(normalize_excel_cell(value) not in (None, "") for value in row_b) else row_a
    labels: list[str] = []

    for col_index, raw_value in enumerate(source_row):
        cell_value = normalize_excel_cell(raw_value)
        if cell_value in (None, ""):
            continue

        header_name = derive_excel_column_name(col_index, header_row_a, header_row_b)
        column_label = header_name or get_column_letter(col_index + 1)
        labels.append(f"{column_label}={truncate_text(cell_value, 28)}")

        if len(labels) == 2:
            break

    if labels:
        return " | ".join(labels)
    return f"linha {row_index}"


def build_excel_focus_label(
    sheet_name: str,
    row_index: int,
    column_letter: str,
    column_name: str | None,
    row_reference: str | None,
) -> str:
    focus = f"planilha \"{sheet_name}\", célula {column_letter}{row_index}"
    if column_name:
        focus += f", coluna \"{column_name}\""
    if row_reference:
        focus += f", registro {row_reference}"
    return focus


def describe_value_profile(value: str | None) -> str | None:
    if value is None:
        return None

    if value == "":
        return "valor vazio"

    value_kind = infer_value_kind(value)
    char_count = len(value)
    line_count = value.count("\n") + 1
    profile = f"{value_kind}, {char_count} caractere(s)"
    if line_count > 1:
        profile += f", {line_count} linha(s)"
    return profile


def infer_value_kind(value: str) -> str:
    lower_value = value.strip().lower()
    if lower_value in {"true", "false", "verdadeiro", "falso"}:
        return "booleano"
    if re.fullmatch(r"-?\d+", value.strip()):
        return "número inteiro"
    if re.fullmatch(r"-?\d+[.,]\d+", value.strip()):
        return "número decimal"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}(t.*)?", lower_value):
        return "data ou data/hora"
    return "texto"


def format_structure_location(path: str) -> str:
    csv_match = re.fullmatch(r"R(\d+)C(\d+)", path)
    if csv_match:
        row_number, column_number = csv_match.groups()
        return f"Linha {row_number}, coluna {column_number}"

    column_count_match = re.fullmatch(r"R(\d+)\.__columns__", path)
    if column_count_match:
        return f"Linha {column_count_match.group(1)}, quantidade de colunas"

    if path.startswith("$"):
        human_path = path.removeprefix("$.").replace("[", " [").replace("].", "] > ").replace(".", " > ")
        return f"Campo JSON {human_path}"

    if path.startswith("/"):
        human_path = path.replace("/@", " > atributo ").replace("/text()", " > texto").replace("/", " > ").strip(" >")
        return f"Elemento XML {human_path}"

    return path


def materialize_upload_to_tempfile(upload: Any) -> str:
    suffix = Path(upload.filename).suffix or ".bin"
    handle = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)

    try:
        upload.stream.seek(0)
        while True:
            chunk = upload.stream.read(1024 * 1024)
            if not chunk:
                break
            handle.write(chunk)
    finally:
        handle.close()
        upload.stream.seek(0)

    return handle.name


def cleanup_temp_path(path: str) -> None:
    try:
        os.remove(path)
    except FileNotFoundError:
        pass


def build_excel_profile_from_path(filename: str, path: str) -> DocumentProfile:
    extension = Path(filename).suffix.lower()
    mime_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    metadata = build_metadata_from_path(path, mime_type, extension)
    return DocumentProfile(
        filename=filename,
        source_type="xlsx",
        format_label="XLSX",
        mode="excel",
        metadata=metadata,
    )


def build_metadata_from_path(path: str, mime_type: str, extension: str) -> dict[str, Any]:
    hasher = hashlib.sha256()

    with open(path, "rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            hasher.update(chunk)

    size_bytes = os.path.getsize(path)
    return {
        "extension": extension or "[sem extensão]",
        "size_bytes": size_bytes,
        "size_human": human_size(size_bytes),
        "mime_type": mime_type,
        "sha256": hasher.hexdigest(),
    }


def load_excel_workbook(content: bytes | str):
    if isinstance(content, str):
        return load_workbook(content, data_only=False, read_only=True)

    return load_workbook(io.BytesIO(content), data_only=False, read_only=True)


def limit_changes(changes: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    if len(changes) <= MAX_DISPLAY_CHANGES:
        return changes, []

    return (
        changes[:MAX_DISPLAY_CHANGES],
        [f"Exibindo apenas as primeiras {MAX_DISPLAY_CHANGES} alterações de {len(changes)}."],
    )


def limit_diff_text(diff_lines: list[str]) -> tuple[str | None, list[str]]:
    if not diff_lines:
        return "Nenhuma diferenca encontrada.", []

    if len(diff_lines) <= MAX_DIFF_LINES:
        return "\n".join(diff_lines), []

    hidden_lines = len(diff_lines) - MAX_DIFF_LINES
    preview = diff_lines[:MAX_DIFF_LINES]
    preview.append(f"... ({hidden_lines} linha(s) de diff ocultas)")
    return "\n".join(preview), [f"Diff unificado limitado a {MAX_DIFF_LINES} linhas para manter a tela leve."]


def build_profile_summary(profile: DocumentProfile) -> dict[str, Any]:
    if profile.mode == "structured":
        content_items = len(profile.records)
    elif profile.mode == "excel":
        content_items = profile.metadata.get("sheet_count")
    elif profile.mode == "text":
        content_items = len(profile.lines)
    elif profile.mode == "archive":
        content_items = len(profile.entries)
    else:
        content_items = None

    return {
        "filename": profile.filename,
        "format_label": profile.format_label,
        "metadata": profile.metadata,
        "content_items": content_items,
        "sheet_names_preview": build_sheet_names_preview(profile.metadata.get("sheet_names")),
    }


def build_summary_items(mode: str, summary: dict[str, Any]) -> list[dict[str, Any]]:
    labels_by_mode = {
        "text": [
            ("differences", "Diferenças encontradas"),
            ("added", "Linhas adicionadas"),
            ("removed", "Linhas removidas"),
            ("changed", "Linhas alteradas"),
            ("unchanged", "Linhas iguais"),
            ("lines_a", "Linhas no arquivo A"),
            ("lines_b", "Linhas no arquivo B"),
            ("similarity", "Semelhança geral"),
        ],
        "structured": [
            ("differences", "Diferenças encontradas"),
            ("added", "Campos adicionados"),
            ("removed", "Campos removidos"),
            ("changed", "Campos alterados"),
            ("unchanged", "Campos iguais"),
            ("fields_a", "Campos no arquivo A"),
            ("fields_b", "Campos no arquivo B"),
            ("similarity", "Semelhança geral"),
        ],
        "excel": [
            ("differences", "Diferenças encontradas"),
            ("changed", "Células alteradas"),
            ("added", "Células adicionadas"),
            ("removed", "Células removidas"),
            ("unchanged", "Células iguais"),
            ("rows_compared", "Linhas comparadas"),
            ("rows_changed", "Linhas com alteração"),
            ("sheets_changed", "Planilhas afetadas"),
            ("sheets_added", "Planilhas adicionadas"),
            ("sheets_removed", "Planilhas removidas"),
            ("sheets_a", "Planilhas no arquivo A"),
            ("sheets_b", "Planilhas no arquivo B"),
            ("similarity", "Semelhança geral"),
        ],
        "archive": [
            ("differences", "Diferenças encontradas"),
            ("added", "Entradas adicionadas"),
            ("removed", "Entradas removidas"),
            ("changed", "Entradas alteradas"),
            ("entries_a", "Entradas no arquivo A"),
            ("entries_b", "Entradas no arquivo B"),
        ],
        "binary": [
            ("difference_ranges", "Faixas divergentes"),
            ("different_bytes", "Bytes diferentes"),
            ("first_difference_offset", "Primeiro byte divergente"),
            ("size_a", "Tamanho do arquivo A"),
            ("size_b", "Tamanho do arquivo B"),
            ("size_delta", "Diferença de tamanho"),
        ],
    }

    summary_items: list[dict[str, Any]] = []
    for key, label in labels_by_mode.get(mode, []):
        if key in summary:
            summary_items.append({"label": label, "value": summary[key]})

    return summary_items


def build_highlight_items(mode: str, summary: dict[str, Any]) -> list[dict[str, Any]]:
    preferred_labels = {
        "text": [
            "Diferenças encontradas",
            "Linhas alteradas",
            "Linhas adicionadas",
            "Semelhança geral",
        ],
        "structured": [
            "Diferenças encontradas",
            "Campos alterados",
            "Campos adicionados",
            "Semelhança geral",
        ],
        "excel": [
            "Diferenças encontradas",
            "Linhas com alteração",
            "Planilhas afetadas",
            "Semelhança geral",
        ],
        "archive": [
            "Diferenças encontradas",
            "Entradas alteradas",
            "Entradas adicionadas",
            "Entradas removidas",
        ],
        "binary": [
            "Faixas divergentes",
            "Bytes diferentes",
            "Primeiro byte divergente",
            "Diferença de tamanho",
        ],
    }

    summary_lookup = {item["label"]: item["value"] for item in build_summary_items(mode, summary)}
    highlights: list[dict[str, Any]] = []
    for label in preferred_labels.get(mode, []):
        if label in summary_lookup:
            highlights.append({"label": label, "value": summary_lookup[label]})

    return highlights[:4]


def build_comparison_overview(
    mode: str,
    summary: dict[str, Any],
    identical: bool,
    changes: list[dict[str, Any]],
) -> dict[str, str]:
    if identical:
        return {
            "title": "Os arquivos são equivalentes dentro da estratégia de comparação aplicada.",
            "body": "Nenhuma diferença relevante foi encontrada no conteúdo analisado.",
        }

    if mode == "excel":
        return {
            "title": (
                f"Foram encontradas {summary.get('differences', 0)} diferença(s) em "
                f"{summary.get('rows_changed', 0)} linha(s) e {summary.get('sheets_changed', 0)} planilha(s)."
            ),
            "body": (
                "A comparação aponta a célula exata afetada, o registro relacionado, o valor anterior, "
                "o valor novo e o contexto da linha para revisão rápida."
            ),
        }

    if mode == "structured":
        return {
            "title": (
                f"Foram encontradas {summary.get('differences', 0)} diferença(s) estruturais em "
                f"{summary.get('changed', 0)} campo(s) alterado(s)."
            ),
            "body": (
                "A leitura destaca o caminho exato do campo, o valor anterior, o valor novo e uma visão "
                "pontual do diff."
            ),
        }

    if mode == "text":
        return {
            "title": (
                f"Foram encontradas {summary.get('differences', 0)} diferença(s) entre "
                f"{summary.get('lines_a', 0)} linha(s) do arquivo A e {summary.get('lines_b', 0)} linha(s) do arquivo B."
            ),
            "body": (
                "A comparação mostra trechos anteriores e novos, além do diff intralinha quando a alteração "
                "cabe em um único ponto."
            ),
        }

    if mode == "archive":
        return {
            "title": (
                f"Foram encontradas {summary.get('differences', 0)} diferença(s) na estrutura interna dos arquivos."
            ),
            "body": "A leitura destaca quais entradas foram adicionadas, removidas ou alteradas dentro do contêiner.",
        }

    first_change = changes[0] if changes else {}
    return {
        "title": (
            f"Foram detectadas {summary.get('difference_ranges', 0)} faixa(s) divergentes e "
            f"{summary.get('different_bytes', 0)} byte(s) diferentes."
        ),
        "body": first_change.get("details", "A comparação binária destaca as primeiras faixas divergentes."),
    }


def build_focus_points(changes: list[dict[str, Any]], limit: int = 3) -> list[dict[str, str]]:
    focus_points: list[dict[str, str]] = []

    for change in changes[:limit]:
        title = change.get("focus_label") or change.get("location_label") or change.get("path") or "Ponto alterado"
        detail = change.get("details") or build_short_transition(change)
        focus_points.append({"title": title, "detail": detail})

    return focus_points


def build_short_transition(change: dict[str, Any]) -> str:
    before_value = change.get("before_value")
    after_value = change.get("after_value")
    if before_value is not None or after_value is not None:
        before_rendered = truncate_text(str(before_value) if before_value is not None else "[sem valor]", 42)
        after_rendered = truncate_text(str(after_value) if after_value is not None else "[sem valor]", 42)
        return f"{before_rendered} -> {after_rendered}"
    return "Veja os detalhes completos no cartão abaixo."


def build_sheet_names_preview(sheet_names: list[str] | None) -> str | None:
    if not sheet_names:
        return None
    if len(sheet_names) <= 4:
        return ", ".join(sheet_names)
    return f"{', '.join(sheet_names[:4])} e mais {len(sheet_names) - 4}"


def build_metadata(content: bytes, mime_type: str, extension: str) -> dict[str, Any]:
    return {
        "extension": extension or "[sem extensão]",
        "size_bytes": len(content),
        "size_human": human_size(len(content)),
        "mime_type": mime_type,
        "sha256": hashlib.sha256(content).hexdigest(),
    }


def build_text_strategy_label(profile_a: DocumentProfile, profile_b: DocumentProfile) -> str:
    if profile_a.mode == "structured" or profile_b.mode == "structured":
        return "Comparação textual entre representações normalizadas"

    if profile_a.source_type == profile_b.source_type:
        if profile_a.source_type == "text":
            return "Comparação textual por linhas com diff intralinha"
        return f"Extração normalizada de {profile_a.format_label} e diff intralinha"

    return f"Comparação textual entre {profile_a.format_label} e {profile_b.format_label}"


def build_structured_strategy_label(profile_a: DocumentProfile, profile_b: DocumentProfile) -> str:
    if profile_a.source_type == profile_b.source_type:
        return f"Comparação por caminho/campo de {profile_a.format_label}"
    return f"Comparação por caminho/campo entre {profile_a.format_label} e {profile_b.format_label}"


def human_size(size_bytes: int) -> str:
    units = ["B", "KB", "MB", "GB", "TB"]
    size = float(size_bytes)

    for unit in units:
        if size < 1024 or unit == units[-1]:
            return f"{size:.1f} {unit}" if unit != "B" else f"{int(size)} {unit}"
        size /= 1024

    return f"{size_bytes} B"


def human_range(start: int, end: int) -> str:
    if start == end:
        return "-"

    first = start + 1
    last = end
    if first == last:
        return f"linha {first}"
    return f"linhas {first}-{last}"


def describe_text_location(i1: int, i2: int, j1: int, j2: int) -> str:
    return f"Arquivo A: {human_range(i1, i2)} | Arquivo B: {human_range(j1, j2)}"


def describe_text_change(tag: str, i1: int, i2: int, j1: int, j2: int) -> str:
    if tag == "insert":
        return f"Conteúdo novo inserido em {human_range(j1, j2)} do segundo arquivo."
    if tag == "delete":
        return f"Conteúdo removido de {human_range(i1, i2)} do primeiro arquivo."
    return (
        f"Trecho alterado entre {human_range(i1, i2)} do primeiro arquivo "
        f"e {human_range(j1, j2)} do segundo arquivo."
    )


def excerpt_lines(lines: list[str], limit: int = 6) -> list[str]:
    if len(lines) <= limit:
        return lines or ["[sem conteúdo]"]

    trimmed = lines[:limit]
    trimmed.append(f"... (+{len(lines) - limit} linhas)")
    return trimmed


def is_textual_profile(profile: DocumentProfile) -> bool:
    return profile.mode in {"text", "structured"}


def profile_to_lines(profile: DocumentProfile) -> list[str]:
    if profile.lines:
        return profile.lines

    if profile.records:
        return render_record_lines(profile.records)

    return []


def render_record_lines(records: dict[str, str]) -> list[str]:
    return [f"{path} = {records[path]}" for path in sorted(records)]


def compute_similarity(before: str, after: str) -> float:
    return difflib.SequenceMatcher(
        a=before,
        b=after,
        autojunk=max(len(before), len(after)) > 20000,
    ).ratio()


def build_matcher(a: list[str], b: list[str]) -> difflib.SequenceMatcher:
    return difflib.SequenceMatcher(
        a=a,
        b=b,
        autojunk=max(len(a), len(b)) > 2000,
    )


def format_percent(value: float) -> str:
    return f"{value * 100:.1f}%"


def build_inline_diff_text(before: str, after: str, *, limit: int = 500) -> str:
    tokens_a = re.findall(r"\s+|\S+", before)
    tokens_b = re.findall(r"\s+|\S+", after)
    matcher = difflib.SequenceMatcher(a=tokens_a, b=tokens_b, autojunk=False)
    parts: list[str] = []

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            parts.append("".join(tokens_a[i1:i2]))
        elif tag == "delete":
            parts.append(f"[-{''.join(tokens_a[i1:i2])}-]")
        elif tag == "insert":
            parts.append(f"{{+{''.join(tokens_b[j1:j2])}+}}")
        else:
            parts.append(f"[-{''.join(tokens_a[i1:i2])}-]")
            parts.append(f"{{+{''.join(tokens_b[j1:j2])}+}}")

    diff_text = "".join(parts)
    if len(diff_text) <= limit:
        return diff_text
    return f"{diff_text[:limit]}... (+{len(diff_text) - limit} caracteres)"


def decode_text_content(content: bytes, allow_latin1: bool = False) -> str | None:
    if not content:
        return ""

    if b"\x00" in content:
        return None

    candidate_encodings = ["utf-8-sig", "utf-8"]
    if allow_latin1:
        candidate_encodings.append("latin-1")

    for encoding in candidate_encodings:
        try:
            decoded = content.decode(encoding)
        except UnicodeDecodeError:
            continue

        if is_probably_text(decoded):
            return decoded

    return None


def is_probably_text(decoded: str) -> bool:
    if not decoded:
        return True

    sample = decoded[:4000]
    weird = sum(1 for char in sample if ord(char) < 32 and char not in "\n\r\t\f")
    return weird / max(len(sample), 1) < 0.05


def normalize_plain_text_lines(text: str) -> list[str]:
    return text.replace("\r\n", "\n").replace("\r", "\n").split("\n")


def normalize_json_records(text: str) -> dict[str, str]:
    parsed = json.loads(text)
    records: dict[str, str] = {}
    flatten_json_value(parsed, "$", records)
    return records or {"$": json.dumps(parsed, ensure_ascii=False, sort_keys=True)}


def flatten_json_value(value: Any, path: str, records: dict[str, str]) -> None:
    if isinstance(value, dict):
        if not value:
            records[path] = "{}"
            return

        for key in sorted(value):
            flatten_json_value(value[key], f"{path}.{key}", records)
        return

    if isinstance(value, list):
        if not value:
            records[path] = "[]"
            return

        for index, item in enumerate(value):
            flatten_json_value(item, f"{path}[{index}]", records)
        return

    records[path] = json.dumps(value, ensure_ascii=False, sort_keys=True)


def normalize_csv_records(text: str) -> dict[str, str]:
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    rows = csv.reader(io.StringIO(text), dialect=dialect)
    records: dict[str, str] = {}

    for row_index, row in enumerate(rows, start=1):
        records[f"R{row_index}.__columns__"] = str(len(row))
        if not row:
            records[f"R{row_index}"] = "[linha vazia]"
            continue

        for col_index, cell in enumerate(row, start=1):
            records[f"R{row_index}C{col_index}"] = cell

    return records or {"R0": "[csv vazio]"}


def normalize_xml_records(text: str) -> dict[str, str]:
    root = ET.fromstring(text)
    records: dict[str, str] = {}
    flatten_xml_element(root, f"/{strip_namespace(root.tag)}[1]", records)
    return records or {"/": "[xml vazio]"}


def flatten_xml_element(element: ET.Element, path: str, records: dict[str, str]) -> None:
    for attr_name, attr_value in sorted(element.attrib.items()):
        records[f"{path}/@{strip_namespace(attr_name)}"] = normalize_space(attr_value)

    direct_text = normalize_space(element.text or "")
    if direct_text:
        records[f"{path}/text()"] = direct_text

    children = [child for child in list(element) if isinstance(child.tag, str)]
    if not children and not element.attrib and not direct_text:
        records[path] = "[vazio]"
        return

    counters: dict[str, int] = {}
    for child in children:
        child_tag = strip_namespace(child.tag)
        counters[child_tag] = counters.get(child_tag, 0) + 1
        flatten_xml_element(child, f"{path}/{child_tag}[{counters[child_tag]}]", records)


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def strip_namespace(tag: str) -> str:
    if "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def looks_like_json(text: str) -> bool:
    stripped = text.strip()
    return bool(stripped) and stripped[0] in "{["


def extract_pdf_lines(content: bytes) -> list[str]:
    reader = PdfReader(io.BytesIO(content))
    lines: list[str] = []

    for page_number, page in enumerate(reader.pages, start=1):
        page_text = (page.extract_text() or "").strip()
        if not page_text:
            lines.append(f"Página {page_number}: [sem texto extraível]")
            continue

        for line_number, line in enumerate(page_text.splitlines(), start=1):
            normalized = normalize_space(line)
            if normalized:
                lines.append(f"Página {page_number}, linha {line_number}: {normalized}")

    return lines or ["[PDF sem texto extraível]"]


def extract_docx_lines(content: bytes) -> list[str]:
    lines: list[str] = []
    paragraph_counts: dict[str, int] = {}
    member_labels = {
        "word/document.xml": "Corpo",
        "word/footnotes.xml": "Nota de rodapé",
        "word/endnotes.xml": "Nota final",
    }

    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        for member_name in archive.namelist():
            if member_name not in member_labels and not member_name.startswith("word/header") and not member_name.startswith("word/footer"):
                continue

            label = member_labels.get(member_name)
            if label is None:
                if member_name.startswith("word/header"):
                    label = "Cabeçalho"
                else:
                    label = "Rodapé"

            xml_root = ET.fromstring(archive.read(member_name))
            for paragraph in xml_root.iterfind(".//w:p", DOCX_NAMESPACES):
                fragments = [node.text for node in paragraph.iterfind(".//w:t", DOCX_NAMESPACES) if node.text]
                if not fragments:
                    continue

                paragraph_counts[label] = paragraph_counts.get(label, 0) + 1
                lines.append(
                    f"{label} paragrafo {paragraph_counts[label]}: {normalize_space(''.join(fragments))}"
                )

    return lines or ["[DOCX sem texto extraível]"]


def extract_xlsx_records(content: bytes) -> dict[str, str]:
    workbook = load_workbook(io.BytesIO(content), data_only=False, read_only=True)
    records: dict[str, str] = {}

    try:
        for worksheet in workbook.worksheets:
            records[f"{worksheet.title}.__sheet__"] = "presente"
            records[f"{worksheet.title}.__extent__"] = f"{worksheet.max_row}x{worksheet.max_column}"

            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    records[f"{worksheet.title}!{cell.coordinate}"] = render_sheet_value(cell.value)
    finally:
        workbook.close()

    return records or {"Workbook": "[sem conteúdo]"}


def render_sheet_value(value: Any) -> str:
    if isinstance(value, str):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def is_zip_archive(content: bytes) -> bool:
    return zipfile.is_zipfile(io.BytesIO(content))


def inspect_zip_entries(content: bytes) -> dict[str, dict[str, Any]]:
    entries: dict[str, dict[str, Any]] = {}

    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        for info in archive.infolist():
            if info.is_dir():
                continue
            entries[info.filename] = {
                "file_size": info.file_size,
                "compressed_size": info.compress_size,
                "crc": f"{info.CRC:08x}",
            }

    return entries
